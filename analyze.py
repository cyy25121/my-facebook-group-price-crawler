#!/usr/bin/env python3
"""解析 data/posts.jsonl 的貼文快照,抽出 (遊戲, 價格),輸出:

  data/listings.csv  — 每筆售價(貼文日期、社團、遊戲、平台、價格、連結、快照時間)
  data/report.xlsx   — Excel 報告:總覽 / 行情統計(公式) / 週走勢 / 明細 / 編輯改價 / 未配對
                       統計工作表使用 Excel 公式,直接改「明細」資料會自動重算

用法: .venv/bin/python analyze.py
"""
import csv
import json
import re
import statistics
import unicodedata
from collections import defaultdict
from datetime import datetime, date as date_cls
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from games import GAME_ALIASES

BASE = Path(__file__).parent
DATA = BASE / "data"

PRICE_RE = re.compile(r"(?<![0-9a-zA-Z.+])([1-9][0-9]{2,4})(?![0-9])")
YEAR_RE = re.compile(r"^20[12][0-9]$")
BUY_HINT = re.compile(r"[收徵](?!藏)|想收|誠收|求售")
SELL_HINT = re.compile(r"[售賣讓]|釋出|出售|脫手")
ITEM_NO_RE = re.compile(r"^\s*[*•·∙▪–—-]*\s*\d{1,3}\s*[\.、,)]\s*")  # 清單項次「1.」「• 101、」
# 這些行的數字不是遊戲價格(運費/周邊/福袋/主機同捆)
SKIP_LINE_RE = re.compile(
    r"運費|郵資|店到店|賣貨便|宅配|郵寄|匯款|福袋|攜帶包|收納|保護殼|保護貼|手把|搖桿|底座|充電|支架|"
    r"記憶卡|amiibo|主機|同捆|套組|組合|合售|合購|整套|全套|包套|限定機|光碟機|加購|"
    r"攜帶盒|卡夾攜帶|卡帶收納|收藏盒|保護包|"
    r"特仕|oled|電力加強|以上|打包|總價|全收")
HARDWARE_RE = re.compile(
    r"主機|手把|搖桿|portal|耳機|psvr|vr2|光碟機|底座|充電|支架|記憶卡|硬碟|ssd|螢幕|攜帶包|收納|保護|擴充")
MAX_GAME_PRICE = 6000  # 超過視為主機/同捆價,不算單片遊戲行情
PLATFORM_TOKENS = [
    ("switch 2", "Switch 2"), ("switch2", "Switch 2"), ("ns2", "Switch 2"),
    ("ns1", "Switch"), ("switch1", "Switch"), ("switch 1", "Switch"),
    ("switch", "Switch"), ("ns ", "Switch"), ("psvita", "PSV"), ("psv", "PSV"),
    ("ps5", "PS5"), ("ps4", "PS4"), ("ps3", "PS3"), ("xbox", "Xbox"),
]


def norm(s):
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"(?<=\d),(?=\d{3})", "", s)  # 千分位逗號 1,100 -> 1100
    return s.lower()


def find_games(text_n):
    """回傳 [(start, end, canonical)],重疊時保留較長的別名。"""
    hits = []
    for canon, aliases in GAME_ALIASES.items():
        for al in aliases:
            start = 0
            while True:
                i = text_n.find(al, start)
                if i < 0:
                    break
                hits.append((i, i + len(al), canon))
                start = i + 1
    hits.sort(key=lambda h: h[1] - h[0], reverse=True)
    kept = []
    for h in hits:
        clash = False
        for idx, k in enumerate(kept):
            overlap = not (h[1] <= k[0] or h[0] >= k[1])
            # 相鄰且中間只有空白/冒號/連字號(「戰神諸神黃昏」「刺客教條:幻象」)→ 同一標題
            gap = text_n[min(h[1], k[1]): max(h[0], k[0])]
            adjacent = gap.strip(" ::-~・·〜.") == "" and len(gap) <= 2
            if overlap or adjacent:
                # 相鄰時視為同一標題:合併 span(避免殘餘短別名再次命中),
                # 若已保留的是系列泛稱(含「(」)而新命中是特定作品 → 採用特定作品名
                if adjacent and not overlap:
                    canon = h[2] if ("(" in k[2] and "(" not in h[2]) else k[2]
                    kept[idx] = (min(h[0], k[0]), max(h[1], k[1]), canon)
                clash = True
                break
        if not clash:
            kept.append(h)
    # 同一行同款遊戲多次命中(「薩爾達傳說」+「王國之淚」拆分命中等)只算一次
    kept.sort()
    seen_canon, uniq = set(), []
    for h in kept:
        if h[2] in seen_canon:
            continue
        seen_canon.add(h[2])
        uniq.append(h)
    return uniq


def find_prices(line_n):
    """從一行文字找價格候選,回傳 [(價格, 是否帶幣值標記)]。"""
    prices = []
    for m in PRICE_RE.finditer(line_n):
        v = m.group(1)
        if YEAR_RE.match(v):
            continue
        if v in ("2077", "1886"):  # 遊戲名稱中的數字(電馭叛客2077、教團1886)
            continue
        # 排除規格(4k/120hz/1tb/825g)、日期、件數
        after = line_n[m.end(): m.end() + 3]
        if re.match(r"k|hz|fps|tb?\b|gb?\b|吋|%", after):
            continue
        if after[:1] in ("年", "月", "日", "片", "台", "件", "人"):
            continue
        n = int(v)
        if 100 <= n <= 99999:
            before = line_n[max(0, m.start() - 2): m.start()]
            marked = bool(re.search(r"[$:¥€:]\s*$|nt\s*$|售\s*$", before)) or after[:1] in ("元", "塊", "$")
            prices.append((n, marked))
    return prices


def pick_price(prices):
    """單一遊戲配多個數字時:優先取帶幣值標記者,否則取最後一個(清單項次通常在前)。"""
    cands = [p for p in prices if p[0] <= MAX_GAME_PRICE]
    if not cands:
        return None
    marked = [p for p in cands if p[1]]
    return (marked[0] if marked else cands[-1])[0]


def detect_platform(text_n, seg="", plat_field="", section=""):
    # 優先序:該行 > 所屬分區標題(如「NS2區」)> 貼文【平台】欄位 > 全文
    for source in (seg, section, plat_field, text_n):
        for token, plat in PLATFORM_TOKENS:
            if token in source:
                return plat
    return ""


def platform_field(text_n):
    """抓【平台】欄位的內容(常見販售文格式)。"""
    m = re.search(r"[【\[]?\s*平台\s*[】\]]?\s*[::]?\s*(.{0,20})", text_n)
    return m.group(1) if m else ""


def extract_listings(text):
    """從貼文內文抽 (game, price, platform, type) 列表。"""
    text_n = norm(text)
    listing_type = "收" if (BUY_HINT.search(text_n) and not SELL_HINT.search(text_n)) else "售"
    out = []
    raw_lines = text_n.split("\n")
    lines = [ITEM_NO_RE.sub("", ln) for ln in raw_lines]  # 去掉清單項次「1.」「101、」
    matched_games_global = find_games(text_n)
    plat_field = platform_field(text_n)

    # 分區標題脈絡:沒有遊戲也沒有價格、但含平台字樣的行(如「---NS2區---」)
    # 之後的行沿用該分區平台,直到下一個分區標題
    sections = [""] * len(lines)
    cur_section = ""
    for i, ln in enumerate(lines):
        if ln.strip() and not find_games(ln) and not find_prices(ln):
            for token, _plat in PLATFORM_TOKENS:
                if token in ln:
                    cur_section = ln
                    break
        sections[i] = cur_section

    def plat(seg="", i=None):
        return detect_platform(text_n, seg, plat_field, sections[i] if i is not None else "")

    # 1) 逐行配對:同一行有遊戲+價格;品名行沒價格時往後看 1~2 行
    used_price_lines = set()
    for i, ln in enumerate(lines):
        if not ln.strip() or SKIP_LINE_RE.search(ln):
            continue
        games = find_games(ln)
        prices = find_prices(ln)
        if games and prices:
            if len(games) == len(prices):
                for (g, (p, _)) in zip([g[2] for g in games], prices):
                    if p <= MAX_GAME_PRICE:
                        out.append((g, p, plat(ln, i), listing_type))
            elif len(games) == 1:
                p = pick_price(prices)
                if p:
                    out.append((games[0][2], p, plat(ln, i), listing_type))
        elif len(games) == 1 and not prices:
            # 品名行,價格在下一行(常見編號清單格式)
            for j in (i + 1, i + 2):
                if j >= len(lines) or j in used_price_lines:
                    continue
                nxt = lines[j]
                if not nxt.strip() or SKIP_LINE_RE.search(nxt) or find_games(nxt):
                    continue
                p = pick_price(find_prices(nxt))
                if p:
                    out.append((games[0][2], p, plat(ln + " " + nxt, i), listing_type))
                    used_price_lines.add(j)
                    break

    # 2) 全文只有一款遊戲:配對最可信的一個價格
    uniq = {g for (_, _, g) in matched_games_global}
    if not out and len(uniq) == 1:
        all_prices = []
        for ln in lines:
            if SKIP_LINE_RE.search(ln):
                continue
            all_prices += find_prices(ln)
        p = pick_price(all_prices)
        if p:
            out.append((next(iter(uniq)), p, plat(), listing_type))

    # 去重(同遊戲同價但不同平台要保留,如 NS1/NS2 版同價分售)
    seen, dedup = set(), []
    for item in out:
        key = (item[0], item[1], item[2])
        if key not in seen:
            seen.add(key)
            dedup.append(item)
    return dedup, bool(matched_games_global)


HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
PRICE_FMT = "#,##0"


def _init_sheet(ws, headers, widths):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.freeze_panes = "A2"


def write_excel(listings, by_game, by_game_plat, edits, unmatched, stats):
    wb = Workbook()
    sell_n = sum(1 for l in listings if l["type"] == "售")

    # ---- 明細(其他工作表的公式都引用這張表)
    det = wb.active
    det.title = "明細"
    _init_sheet(det, ["日期", "週", "社團", "post_id", "作者", "類型", "平台", "遊戲", "價格", "連結", "抓取批次"],
                [11, 5, 26, 19, 14, 6, 9, 24, 9, 46, 12])
    rows = sorted(listings, key=lambda r: r["date"])  # 日期升冪,供 LOOKUP 取最新一筆
    for i, l in enumerate(rows, start=2):
        d = datetime.strptime(l["date"], "%Y-%m-%d").date()
        det.cell(row=i, column=1, value=d).number_format = "yyyy-mm-dd"
        det.cell(row=i, column=2, value=f"=_xlfn.ISOWEEKNUM(A{i})")
        det.cell(row=i, column=3, value=l["group"])
        det.cell(row=i, column=4, value=l["post_id"])
        det.cell(row=i, column=5, value=l["author"])
        det.cell(row=i, column=6, value=l["type"])
        det.cell(row=i, column=7, value=l["platform"] or "?")
        det.cell(row=i, column=8, value=l["game"])
        det.cell(row=i, column=9, value=l["price"]).number_format = PRICE_FMT
        c = det.cell(row=i, column=10, value=l["url"])
        c.hyperlink, c.font = l["url"], LINK_FONT
        det.cell(row=i, column=11, value=l["scraped_at"][:10])
        for col in (1, 2, 3, 4, 5, 6, 7, 8, 11):
            det.cell(row=i, column=col).font = BODY_FONT
    n = len(rows) + 1

    def cond(r, week_ref=None, plat=True):
        c = f'(明細!$H$2:$H${n}=$A{r})*(明細!$F$2:$F${n}="售")'
        if plat:
            c += f"*(明細!$G$2:$G${n}=$B{r})"
        if week_ref:
            c += f"*(明細!$B$2:$B${n}={week_ref})"
        return c

    # ---- 行情統計(遊戲 × 平台,公式計算)
    st = wb.create_sheet("行情統計")
    _init_sheet(st, ["遊戲", "平台", "筆數", "最低", "中位", "最高", "最新價", "最新日期"],
                [24, 9, 7, 8, 8, 8, 8, 11])
    order = sorted(by_game_plat.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    for i, ((g, plat), _ls) in enumerate(order, start=2):
        st.cell(row=i, column=1, value=g)
        st.cell(row=i, column=2, value=plat)
        crit = f'明細!$H$2:$H${n},$A{i},明細!$G$2:$G${n},$B{i},明細!$F$2:$F${n},"售"'
        st.cell(row=i, column=3, value=f"=COUNTIFS({crit})")
        st.cell(row=i, column=4, value=f"=_xlfn.MINIFS(明細!$I$2:$I${n},{crit})")
        st.cell(row=i, column=5, value=ArrayFormula(f"E{i}", f"=MEDIAN(IF({cond(i)},明細!$I$2:$I${n}))"))
        st.cell(row=i, column=6, value=f"=_xlfn.MAXIFS(明細!$I$2:$I${n},{crit})")
        st.cell(row=i, column=7, value=f"=LOOKUP(2,1/({cond(i)}),明細!$I$2:$I${n})")
        st.cell(row=i, column=8, value=f"=LOOKUP(2,1/({cond(i)}),明細!$A$2:$A${n})")
        for col in range(1, 9):
            st.cell(row=i, column=col).font = BODY_FONT
        for col in (4, 5, 6, 7):
            st.cell(row=i, column=col).number_format = PRICE_FMT
        st.cell(row=i, column=8).number_format = "yyyy-mm-dd"

    # ---- 週走勢(遊戲 × 平台,每週中位價)
    weeks = sorted({date_cls.fromisoformat(l["date"]).isocalendar()[1]
                    for l in listings if l["type"] == "售"})
    tr = wb.create_sheet("週走勢")
    _init_sheet(tr, ["遊戲(週中位價)", "平台"] + [f"W{w}" for w in weeks], [24, 9] + [9] * len(weeks))
    top_pairs = [(g, p) for (g, p), ls in sorted(by_game_plat.items(), key=lambda kv: (-len(kv[1]), kv[0]))
                 if len(ls) >= 5]
    for i, (g, plat) in enumerate(top_pairs, start=2):
        tr.cell(row=i, column=1, value=g).font = BODY_FONT
        tr.cell(row=i, column=2, value=plat).font = BODY_FONT
        for j, w in enumerate(weeks, start=3):
            ref = f"{get_column_letter(j)}{i}"
            f = f'=IFERROR(MEDIAN(IF({cond(i, week_ref=w)},明細!$I$2:$I${n})),"")'
            cell = tr.cell(row=i, column=j, value=ArrayFormula(ref, f))
            cell.font, cell.number_format = BODY_FONT, PRICE_FMT

    # ---- 編輯改價
    ed = wb.create_sheet("編輯改價")
    _init_sheet(ed, ["post_id", "遊戲", "原價", "改後", "發現於", "連結"], [19, 24, 8, 8, 11, 46])
    if edits:
        for i, (pid, url, g, p0, p1, at) in enumerate(edits, start=2):
            ed.append([pid, g, p0, p1, at[:10], url])
            ed.cell(row=i, column=6).hyperlink = url
            ed.cell(row=i, column=6).font = LINK_FONT
    else:
        ed["A2"] = "(目前無偵測到編輯改價;重跑 crawler.py 累積快照後即可比對)"
        ed["A2"].font = BODY_FONT

    # ---- 未配對(供擴充 games.py)
    um = wb.create_sheet("未配對")
    _init_sheet(um, ["發文時間", "連結", "內文(前300字)"], [17, 46, 100])
    for i, r in enumerate(unmatched, start=2):
        um.cell(row=i, column=1, value=r["created_at"][:16].replace("T", " ")).font = BODY_FONT
        c = um.cell(row=i, column=2, value=r["url"])
        c.hyperlink, c.font = r["url"], LINK_FONT
        t = um.cell(row=i, column=3, value=(r.get("text") or "")[:300].replace("\n", " / "))
        t.font = BODY_FONT
        t.alignment = Alignment(wrap_text=False)

    # ---- 總覽
    ov = wb.create_sheet("總覽", 0)
    ov.column_dimensions["A"].width = 30
    ov.column_dimensions["B"].width = 18
    items = [
        ("遊戲片價格統計報告", None),
        ("產生時間", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("貼文快照(筆)", stats["snapshots"]),
        ("不重複貼文(篇)", stats["posts"]),
        ("解析出售價(筆)", '=COUNTIF(明細!F:F,"售")'),
        ("收購文(筆)", '=COUNTIF(明細!F:F,"收")'),
        ("跨團轉貼去重(筆)", stats["reposts"]),
        ("遊戲 × 平台組合(項)", "=COUNTA(行情統計!A:A)-1"),
        ("無法配對遊戲(篇)", stats["unmatched"]),
        ("主機/周邊文已排除(篇)", stats["hardware"]),
    ]
    for i, (k, v) in enumerate(items, start=1):
        ov.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=(i == 1), size=12 if i == 1 else 10)
        if v is not None:
            ov.cell(row=i, column=2, value=v).font = BODY_FONT

    wb.save(DATA / "report.xlsx")


def main():
    src = DATA / "posts.jsonl"
    if not src.exists():
        raise SystemExit("找不到 data/posts.jsonl,請先執行 crawler.py")

    snapshots = [json.loads(l) for l in src.read_text(encoding="utf-8").splitlines() if l.strip()]
    # 同一篇貼文同一次執行只留一筆;不同執行批次都保留(偵測編輯)
    by_post = defaultdict(dict)   # post_id -> {scraped_at: rec}
    for r in snapshots:
        by_post[r["post_id"]][r["scraped_at"]] = r

    listings = []
    unmatched = []
    hardware = 0
    edits = []

    for pid, runs in by_post.items():
        ordered = sorted(runs.values(), key=lambda r: r["scraped_at"])
        latest = ordered[-1]
        lst, had_game = extract_listings(latest.get("text") or "")
        for (g, price, plat, typ) in lst:
            listings.append({
                "date": latest["created_at"][:10],
                "group": latest["group_name"],
                "post_id": pid,
                "author": latest.get("author") or "",
                "type": typ,
                "platform": plat,
                "game": g,
                "price": price,
                "url": latest["url"],
                "scraped_at": latest["scraped_at"],
            })
        if not lst and (latest.get("text") or "").strip():
            text_n = norm(latest["text"])
            any_price = any(find_prices(l) for l in text_n.split("\n"))
            if any_price:
                # 主機/周邊文不算「對應不到的遊戲」
                if HARDWARE_RE.search(text_n) and not find_games(text_n):
                    hardware += 1
                else:
                    unmatched.append(latest)

        # 編輯偵測:跨執行批次內容有變 → 比對價格
        if len(ordered) > 1:
            prev_map = None
            for snap in ordered:
                cur, _ = extract_listings(snap.get("text") or "")
                cur_map = {g: p for (g, p, _, _) in cur}
                if prev_map is not None:
                    for g, p in cur_map.items():
                        if g in prev_map and prev_map[g] != p:
                            edits.append((pid, snap["url"], g, prev_map[g], p, snap["scraped_at"]))
                prev_map = cur_map

    # 同作者同日同遊戲同價的跨團轉貼(他團同步)只留一筆
    seen_repost = set()
    deduped = []
    for l in sorted(listings, key=lambda r: r["scraped_at"]):
        key = (l["author"], l["game"], l["price"], l["platform"], l["date"], l["type"])
        if key in seen_repost:
            continue
        seen_repost.add(key)
        deduped.append(l)
    reposts = len(listings) - len(deduped)
    listings = deduped

    # ---- listings.csv
    DATA.mkdir(exist_ok=True)
    with (DATA / "listings.csv").open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["date", "group", "post_id", "author", "type", "platform", "game", "price", "url", "scraped_at"])
        w.writeheader()
        for row in sorted(listings, key=lambda r: (r["game"], r["date"])):
            w.writerow(row)

    # ---- report.xlsx
    sell = [l for l in listings if l["type"] == "售"]
    by_game = defaultdict(list)          # 遊戲 -> listings(跨平台,給走勢用)
    by_game_plat = defaultdict(list)     # (遊戲, 平台) -> listings
    for l in sell:
        by_game[l["game"]].append(l)
        by_game_plat[(l["game"], l["platform"] or "?")].append(l)

    stats = {"snapshots": len(snapshots), "posts": len(by_post), "reposts": reposts,
             "unmatched": len(unmatched), "hardware": hardware}
    write_excel(listings, by_game, by_game_plat, edits, unmatched, stats)

    print(f"[analyze] 售價 {len(sell)} 筆 / 遊戲 {len(by_game)} 款 / 編輯改價 {len(edits)} 筆")
    print(f"[analyze] 輸出 → data/listings.csv, data/report.xlsx")


if __name__ == "__main__":
    main()
